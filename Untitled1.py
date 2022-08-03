#!/usr/bin/env python
# coding: utf-8

# In[1]:


#import copy, sys
from PyPDF2 import PdfFileWriter, PdfFileReader
import os
#from PIL import Image
#from pdf2image import convert_from_path
#import pytesseract
from openpyxl import Workbook,load_workbook
import glob
import PySimpleGUI as sg                        # Part 1 - The import

# Define the window's contents
layout = [  [sg.Text("excell dosya adı")],     # Part 2 - The Layout
            [sg.Input("icra excell deneme")],
            #[sg.Text("dosya adı")],     # Part 2 - The Layout
            #[sg.Input()],
            [sg.Button('Ok')] ]

# Create the window
window = sg.Window('Window Title', layout)      # Part 3 - Window Defintion

# Display and interact with the Window
event, values = window.read()                   # Part 4 - Event loop or Window.read call

# Do something with the information gathered
#print('Hello', values[0], "! Thanks for trying PySimpleGUI")
#print('Hello', values[1], "! Thanks for trying PySimpleGUI")
# Finish up by removing from the screen
window.close()



#input2 = PdfFileReader(values[0]) 
output2=PdfFileWriter()    
#pytesseract.pytesseract.tesseract_cmd = r'Tesseract-OCR\tesseract'

wb = load_workbook(str(values[0])+".xlsx")
ws = wb.active
EsasNo=list()
for row in range(2,ws.max_row+1):
    cell_name = "{}{}".format("E", row)
    ws[cell_name].value
    EsasNo.append((cell_name,ws[cell_name].value))
print(glob.glob("*.pdf"))
for x in range(len(glob.glob("*.pdf"))):
    
    pdfDoc= PdfFileReader(glob.glob("*.pdf")[x])
    for y in range(len(pdfDoc.pages)):
        currentPage = pdfDoc.getPage(y)
        txt=currentPage.extractText()
        txt = txt.split()
        
        #print("--------------------txt"+ str(x)+str(y))
        #print(txt)
        #print("--------------------txt"+ str(x)+str(y))

#pageOne = input2.getPage(0)
#txt=pageOne.extractText()

#txt=txt.split("\n")

        counter= 0
        count3=0
        count2=0
        pageTupleList=list()
        dataList = list()
        a=""
        b=""
        c=""
        control=0
        for i in txt:
            
            if r"/" in i and control == 0:
                a=i
                control=1
            if "TL" in i and control == 1:
                b=c
                pageTupleList.append((a,b[11:]))
                control=0
            c=i 
        print(pageTupleList)
            

            
       # print(pageTupleList)
        print(glob.glob("*.pdf")[x])
        print("-------------"+ str(x)+"    "+str(y))

        Gcounter=1
        for i in pageTupleList:
                
                for e in EsasNo:
                    if(e[1]==i[0]):
                        ws["G"+e[0][1:]] = i[1]
                        print("G"+e[0][1:] +"  "+ i[1])
                Gcounter+=1
wb.save(str(values[0])+".xlsx")


#newPdf= PdfFileReader("2022 HAZİRAN KAFEM-TRAFİK İCRA LİSTESİ .pdf");
#txt= newPdf.getPage(0).extractText()
#txt=txt.split("\n")
#print(txt)
#newPdf= PdfFileReader("FEN İŞLERİ HAZİRAN NAFAKA LİSTESİ.pdf");
#txt= newPdf.getPage(0).extractText()
#txt=txt.split("\n")
#print(txt)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




